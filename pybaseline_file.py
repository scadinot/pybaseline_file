import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from pybaselines.whittaker import aspls
from scipy.signal import savgol_filter
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import os
import glob
from tkinter import IntVar, Radiobutton, Tk, filedialog, Button, Label, Frame, StringVar, messagebox, ttk, Text, OptionMenu
from multiprocessing import freeze_support
import re
import time
import platform
import subprocess

def open_folder(path):
    if platform.system() == "Windows":
        os.startfile(path)
    elif platform.system() == "Darwin":  # macOS
        subprocess.call(["open", path])
    else:  # Linux
        subprocess.call(["xdg-open", path])

def readFile(filePath, sep, decimal) -> (pd.DataFrame|None):
    with open(filePath, encoding="latin1") as fileStream:
        dataFrame = pd.read_csv(fileStream, sep=sep, skiprows=1, usecols=[0, 1], names=["Potential", "Current"], decimal=decimal)
    return dataFrame

def processData(dataFrame) -> tuple:
    dataFrame = dataFrame[dataFrame["Current"] != 0].sort_values("Potential").reset_index(drop=True)
    potentialValues = dataFrame["Potential"].values
    signalValues = -dataFrame["Current"].values  # Inversion du courant
    return potentialValues, signalValues, dataFrame

def smoothSignal(signalValues) -> np.ndarray:
    return savgol_filter(signalValues, window_length=11, polyorder=2)

def getPeakValue(signalValues, potentialValues, marginRatio=0.10, maxSlope=None) -> tuple:
    n = len(signalValues)
    margin = int(n * marginRatio)
    searchRegion = signalValues[margin:-margin]
    potentialsRegion = potentialValues[margin:-margin]

    if maxSlope is not None:
        slopes = np.gradient(searchRegion, potentialsRegion)
        validIndices = np.where(np.abs(slopes) < maxSlope)[0]
        if len(validIndices) == 0:
            return potentialValues[margin], signalValues[margin]
        bestIndex = validIndices[np.argmax(searchRegion[validIndices])]
        index = bestIndex + margin
    else:
        indexInRegion = np.argmax(searchRegion)
        index = indexInRegion + margin

    return potentialValues[index], signalValues[index]

def calculateSignalBaseLine(signalValues, potentialValues, xPeakVoltage, exclusionWidthRatio=0.03, lambdaFactor=1e3) -> tuple[np.ndarray, tuple[float, float]]:
    n = len(signalValues)
    lam = lambdaFactor * (n ** 2)
    exclusionWidth = exclusionWidthRatio * (potentialValues[-1] - potentialValues[0])
    weights = np.ones_like(potentialValues)
    exclusion_min = xPeakVoltage - exclusionWidth
    exclusion_max = xPeakVoltage + exclusionWidth
    weights[(potentialValues > exclusion_min) & (potentialValues < exclusion_max)] = 0.001
    baselineValues, _ = aspls(signalValues, lam=lam, diff_order=2, weights=weights, tol=1e-2, max_iter=25)
    return baselineValues, (exclusion_min, exclusion_max)

def plotSignalAnalysis(potentialValues, signalValues, signalSmoothed, baseline, signalCorrected, xCorrectedVoltage, yCorrectedCurrent, fileName, outputFolder) -> None:
    plt.figure(figsize=(10, 6))
    plt.plot(potentialValues, signalValues, label="Signal brut", alpha=0.5)
    plt.plot(potentialValues, signalSmoothed, label="Signal lissé", linewidth=2)
    plt.plot(potentialValues, baseline, label="Baseline estimée (asPLS)", linestyle='--')
    plt.plot(potentialValues, signalCorrected, label="Signal corrigé", linewidth=3)
    plt.plot(xCorrectedVoltage, yCorrectedCurrent, 'mo', label=f"Pic corrigé à {xCorrectedVoltage:.3f} V ({yCorrectedCurrent*1e3:.3f} mA)")
    plt.axvline(xCorrectedVoltage, color='magenta', linestyle=':', linewidth=1)
    plt.xlabel("Potentiel (V)")
    plt.ylabel("Courant (A)")
    plt.title(f"Correction de baseline : {fileName}")
    plt.legend()
    plt.grid(True)
    plt.tight_layout()
    outputPath = os.path.join(outputFolder, fileName.replace(".txt", ".png"))
    plt.savefig(outputPath, dpi=300, bbox_inches='tight')
    plt.close()

def processSignalFile(filePath, outputFolder, sep, decimal, export_choice) -> dict:
    try:
        fileName = os.path.basename(filePath)
        dataFrame = readFile(filePath, sep=sep, decimal=decimal)
        if dataFrame is None:
            return None

        potentialValues, signalValues, cleaned_df = processData(dataFrame)
        signalSmoothed = smoothSignal(signalValues)
        xPeakVoltage, yPeakCurrent = getPeakValue(signalSmoothed, potentialValues, marginRatio=0.10, maxSlope=500)
        baseline, _ = calculateSignalBaseLine(signalSmoothed, potentialValues, xPeakVoltage, exclusionWidthRatio=0.03, lambdaFactor=1e3)
        signalCorrected = signalSmoothed - baseline
        xCorrectedVoltage, yCorrectedCurrent = getPeakValue(signalCorrected, potentialValues, marginRatio=0.10, maxSlope=500)
        plotSignalAnalysis(potentialValues, signalValues, signalSmoothed, baseline, signalCorrected, xCorrectedVoltage, yCorrectedCurrent, fileName, outputFolder)

        if export_choice == 1:
            cleaned_df.to_csv(os.path.join(outputFolder, fileName.replace(".txt", ".csv")), index=False)
        elif export_choice == 2:
            cleaned_df.to_excel(os.path.join(outputFolder, fileName.replace(".txt", ".xlsx")), index=False)

        match = re.match(r"(.+)_C(\d{2})\.txt", fileName)
        baseName = match.group(1) if match else fileName
        electrode = f"C{match.group(2)}" if match else ""

        return {
            "Base": baseName,
            f"{electrode} - Tension (V)": xCorrectedVoltage,
            f"{electrode} - Courant (A)": yCorrectedCurrent,
            f"{electrode} - Charge (C)": "",
        }

    except Exception as exception:
        print(f"Erreur lors de la lecture de {fileName} : {exception}")
        return {"error": f"Erreur dans le fichier {fileName} : {str(exception)}"}

def main():
    freeze_support()
    launch_gui()

def launch_gui():
    def select_folder():
        path = filedialog.askdirectory(title="Sélectionnez le dossier contenant les fichiers .txt")
        if path:
            folder_path.set(path)

    def run_analysis():
        export_choice = export_option.get()
        #export_csv = export_choice == 1
        #export_excel = export_choice == 2

        log_box.config(state="normal")
        log_box.delete("1.0", "end")
        log_box.config(state="disabled")
        inputFolder = folder_path.get()
        if not inputFolder or not os.path.isdir(inputFolder):
            messagebox.showerror("Erreur", "Veuillez sélectionner un dossier valide.")
            return

        sep_label = sep_var.get()
        sep_map = {"Tabulation": "\t", "Virgule": ",", "Point-virgule": ";", "Espace": " "}
        sep = sep_map.get(sep_label, "\t")
        decimal_label = decimal_var.get()
        decimal_map = {"Point": ".", "Virgule": ","}
        decimal = decimal_map.get(decimal_label, ".")

        folderName = os.path.basename(os.path.normpath(inputFolder))
        outputFolder = os.path.join(os.path.dirname(inputFolder), folderName + " (results)")
        os.makedirs(outputFolder, exist_ok=True)

        # Nettoyage du dossier de sortie
        log_box.config(state="normal")
        log_box.insert("end", "Nettoyage du dossier de sortie...\n")
        log_box.config(state="disabled")
        for file in glob.glob(os.path.join(outputFolder, "*")):
            if file.endswith((".png", ".csv", ".xlsx")):
                os.remove(file)

        filePaths = sorted(glob.glob(os.path.join(inputFolder, "*.txt")))
        fileProcessingArgs = [(filePath, outputFolder, sep, decimal, export_choice) for filePath in filePaths]

        results = []
        start_time = time.time()

        progress_bar["maximum"] = len(filePaths)
        progress_bar["value"] = 0

        for i, filePath in enumerate(filePaths):
            result = processSignalFile(filePath, outputFolder, sep, decimal, export_choice)
            log_box.config(state="normal")
            if result:
                if "error" in result:
                        log_box.insert("end", f"Erreur : {result['error']}\n", ("error",))
                else:
                    results.append(result)
                    log_box.insert("end", f"Traitement : {os.path.basename(filePath)}\n")
            else:
                log_box.insert("end", f"Fichier ignoré ou invalide : {os.path.basename(filePath)}\n")

            log_box.update_idletasks()
            log_box.see("end")
            log_box.tag_config("error", foreground="red")
            log_box.config(state="disabled")
            progress_bar["value"] = i + 1
            root.update_idletasks()

        if results:
            df = pd.DataFrame(results)
            df = df.groupby("Base").first().reset_index()
            df.insert(1, 'Fréq (Hz)', 50.0)
            excel_path = os.path.join(outputFolder, folderName + ".xlsx")
            df.to_excel(excel_path, index=False)

            wb = load_workbook(excel_path)
            ws = wb.active
            header = [cell.value for cell in ws[1]]
            freq_col_letter = get_column_letter(header.index('Fréq (Hz)') + 1)

            for col_index, col_name in enumerate(header):
                if col_name.endswith("- Courant (A)"):
                    elec = col_name.split(" - ")[0]
                    charge_col = f"{elec} - Charge (C)"
                    if charge_col in header:
                        charge_col_index = header.index(charge_col) + 1
                        current_col_letter = get_column_letter(col_index + 1)
                        for row in range(2, ws.max_row + 1):
                            formula = f"={current_col_letter}{row}/{freq_col_letter}{row}"
                            ws.cell(row=row, column=charge_col_index, value=formula)

            wb.save(excel_path)
            log_box.config(state="normal")
            duration = time.time() - start_time
            summary = f"\nTraitement terminé avec succès.\nFichiers traités : {len(results)} / {len(filePaths)}\nTemps écoulé : {duration:.2f} secondes.\n\n"
            log_box.insert("end", summary)
            log_box.update_idletasks()
            log_box.see("end")
            log_box.config(state="disabled")
            messagebox.showinfo("Succès", "Traitement terminé avec succès.")
            result_button.config(state="normal")

    root = Tk()
    root.resizable(True, True)
    
    root.title("Analyse de fichiers SWV")
    root.geometry("700x400")
    root.minsize(600, 400)

    folder_path = StringVar()
    sep_options = ["Tabulation", "Virgule", "Point-virgule", "Espace"]
    decimal_options = ["Point", "Virgule"]

    sep_var = StringVar(value="Tabulation")
    decimal_var = StringVar(value="Point")
    export_option = IntVar(value=0)

    main_frame = Frame(root, padx=10, pady=10)
    main_frame.grid(row=0, column=0, sticky="nsew")
    main_frame.grid_columnconfigure(1, weight=1)
    root.grid_rowconfigure(0, weight=1)
    root.grid_columnconfigure(0, weight=1)

    Label(main_frame, text="Dossier d'entrée :").grid(row=0, column=0, sticky="w")
    Label(main_frame, textvariable=folder_path, relief="sunken", anchor="w", width=50).grid(row=0, column=1, padx=5, sticky="ew")
    Button(main_frame, text="Parcourir", command=select_folder).grid(row=0, column=2, padx=5)

    settings_frame = ttk.LabelFrame(main_frame, text="Paramètres de lecture")
    settings_frame.grid(row=1, column=0, columnspan=3, pady=(10, 5), sticky="ew")

    Label(settings_frame, text="Séparateur de colonnes :").grid(row=0, column=0, sticky="w")
    sep_radio_frame = Frame(settings_frame)
    sep_radio_frame.grid(row=0, column=1, columnspan=4, sticky="w")
    for i, txt in enumerate(sep_options):
        ttk.Radiobutton(sep_radio_frame, text=txt, variable=sep_var, value=txt).grid(row=0, column=i, sticky="w", padx=(0, 10))

    Label(settings_frame, text="Séparateur décimal :").grid(row=1, column=0, sticky="w")
    dec_radio_frame = Frame(settings_frame)
    dec_radio_frame.grid(row=1, column=1, columnspan=4, sticky="w")
    for i, txt in enumerate(decimal_options):
        ttk.Radiobutton(dec_radio_frame, text=txt, variable=decimal_var, value=txt).grid(row=0, column=i, sticky="w", padx=(0, 10))

    Label(settings_frame, text="Export des fichiers :").grid(row=2, column=0, sticky="w", pady=(5, 0))
    export_radio_frame = Frame(settings_frame)
    export_radio_frame.grid(row=2, column=1, columnspan=4, sticky="w")
    Radiobutton(export_radio_frame, text="Ne pas exporter", variable=export_option, value=0).pack(side="left", padx=(0, 10))
    Radiobutton(export_radio_frame, text="Exporter au format .CSV", variable=export_option, value=1).pack(side="left", padx=(0, 10))
    Radiobutton(export_radio_frame, text="Exporter au format Excel", variable=export_option, value=2).pack(side="left")

    progress_frame = ttk.LabelFrame(main_frame, text="Progression du traitement")
    progress_frame.grid(row=2, column=0, columnspan=3, sticky="ew", padx=2, pady=(5, 5))
    progress_bar = ttk.Progressbar(progress_frame, orient="horizontal", mode="determinate")
    progress_bar.pack(fill="x", padx=5, pady=5)

    log_frame = ttk.LabelFrame(main_frame, text="Journal de traitement")
    log_frame.grid(row=3, column=0, columnspan=3, sticky="nsew", padx=2, pady=(0, 5))
    main_frame.grid_rowconfigure(3, weight=1)
    log_box = Text(log_frame, relief="sunken", wrap="word", height=10, bg="white")
    log_box.pack(expand=True, fill="both", padx=5, pady=5)
    log_box.config(state="disabled")

    action_frame = Frame(main_frame)
    action_frame.grid(row=4, column=0, columnspan=3, sticky="ew")
    Button(action_frame, text="Lancer l'analyse", command=run_analysis).pack(side="right", padx=5, pady=5)
    result_button = Button(action_frame, text="Ouvrir le dossier de résultats", state="disabled", command=lambda: open_folder(folder_path.get() + " (results)"))
    result_button.pack(side="right", padx=5, pady=5)

    root.mainloop()

if __name__ == '__main__':
    main()
